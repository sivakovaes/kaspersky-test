from openpyxl import Workbook
from datetime import datetime


def process_xlsx(**kwargs):
    date = datetime.now().strftime("%Y-%m-%d")
    time = datetime.now().strftime("%H-%M-%S")

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Values'
    ws['A3'] = 'Bitcoin'
    ws['A4'] = 'Etherium'
    ws['A5'] = 'Litecoin'
    ws['B3'] = kwargs.get('BTC')
    ws['B4'] = kwargs.get('ETH')
    ws['B5'] = kwargs.get('LTC')

    wb.save(f"currency_{date}_{time}.xlsx")
